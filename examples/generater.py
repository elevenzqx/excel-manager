from  openpyxl import load_workbook
import re

# 定义一个类，用于存放表格的累计数据，如有需要可以加入其它要统计的数据
class TableItem:
    # 初始化类对象，传入两个数据
    def __init__(self, showCount, clickCount):
        self.showSum = int(showCount)
        self.clickSum = int(clickCount)
        self.totalSum = int(1)
    # 增加展示量的方法
    def updateAndIncrCount(self, showCount, clickCount):
        self.showSum += showCount
        self.totalSum += 1
        self.clickSum += clickCount
    def getClickRate(self):
        if self.showSum <= 0:
            return 0
        return float(self.clickSum) / self.showSum
    def getShowSum(self):
        return self.showSum
    def getClickSum(self):
        return self.clickSum
    def getTotalLine(self):
        return self.totalSum

# 字典表，可以新增需要的字典数据
dictionary = {
    "woman": "women",
    "suits": "suit",
    "sweats": "sweat",
    "for women": "women"
}

# 替换字典表中存在的词
def replaceTableValue( value ):
    returnVal = value
    for key in dictionary:
        returnVal = returnVal.replace(key, dictionary[key])
    return returnVal

# 替换字典表中存在的词
def getActivityName( value ):
    pattern = re.compile(r'(([A-Za-z])[A-Za-z0-9]+)')
    m = pattern.search(value)
    if m:
        return m.group(1)
    return ""

# 加载表格数据
wb = load_workbook('data.xlsx')
# 加载当前的表格sheet
ws3 = wb.get_sheet_by_name("sheet")

i = 2
tableKey = ws3["I{}".format(i)].value

# for循环遍历所有的行
## 定义字典结构，用于存放待输出的数据
### map[key] = value, key为输出之后的单数标准值， value为表格对应行数据的累加值
### map[key] = value, key为输出之后的单数标准值， value为表格对应行数据的累加值
mydictTotal = dict()
myDict = None
while tableKey is not None:
    activityName = getActivityName(ws3["E{}".format(i)].value)
    if mydictTotal.get(activityName):
        myDict = mydictTotal.get(activityName)
    else:
        myDict = dict()
        mydictTotal[activityName] = myDict
    
    formatKey = replaceTableValue(tableKey)
    ### 1. if key in map: 把当前行的数据累加到value的不同字段中
    if myDict.get(formatKey):
        rowData = myDict.get(formatKey)
        rowData.updateAndIncrCount(int(ws3["J{}".format(i)].value), int(ws3["K{}".format(i)].value))
    else:
        showCount = int(ws3["J{}".format(i)].value)
        clickCount = int(ws3["K{}".format(i)].value)
        ### 2. if key not in map: 创建TableItem对象用于存放当前的点击数据
        rowData = TableItem(showCount, clickCount)
        ### 3. 添加到字典中
        myDict[formatKey] = rowData
    i+=1
    tableKey = ws3["I{}".format(i)].value

wsout = wb.create_sheet("output")
wsout.append(["序列", "搜索词", "展现总量", "点击量", "点击率", "总数量"])
# 遍历结果，转出最终表格
for key in mydictTotal.keys():
    myDict = mydictTotal.get(key)
    for subKey in myDict:
        wsout.append([key, subKey, myDict[subKey].getShowSum(), myDict[subKey].getClickSum(), "{:.2f}%".format(myDict[subKey].getClickRate() * 100), myDict[subKey].getTotalLine()])
        ## 这个地方换成表格的输出即可
        # print("序列: {}, 搜索词: {}, 展现量总: {}, 点击量总: {}, 点击率: {:.2f}%, 总数量: {}".format(key,
        #     subKey, myDict[subKey].getShowSum(), myDict[subKey].getClickSum(), myDict[subKey].getClickRate(), myDict[subKey].getTotalLine()))

wb.save("dataOut.xlsx")

# print(ws3) 
# print(ws3['I2'].value)

# 声明一个函数，用于将对应的I列的值进行转换，转换成单数的形式，示例：xxx for xxx womens -> xxx xxx women
# escap = ["for", "with"]

# print(dictionary["woman"])

# line = ws3["I2"].value

# print(replaceTableValue("this is for woman"))
# pattern = re.compile("([a-z]+)(s|'s)")

# def formatValue( value ):
#     str = value
#     m = pattern.match(str)
#     if m:
#         str = m.group(1)

#     if str in dictionary:
#         return dictionary[str]
#     return str


# print(formatValue(u"women's"))
# print(formatValue(u"womens"))
# print(formatValue(u"women"))
# print(formatValue(u"woman"))

# # print(replaceValue("this is womens"))
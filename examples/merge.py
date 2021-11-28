from openpyxl import *

def read( value, output ):
    wb = load_workbook(value)
    ws = wb.active
    maxRow = ws.max_row
    i = 2
    while i < maxRow:
        row_list = []
        for t in ws[i]:
            row_list.append(t.value)
        output.append(row_list)
        i += 1

wout = Workbook()
output = wout.active

output.title = "汇总表"
output.append(["投放","搜索词"])
read('./data/data1.xlsx', output)
read('./data/data2.xlsx', output)

wout.save("./target/output.xlsx")
